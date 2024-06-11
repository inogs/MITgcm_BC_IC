from openpyxl import load_workbook
from pathlib import Path
import numpy as np

import os
import sys

import rivers


xlsfile = os.getenv("RIVERDATA")
if xlsfile is None:
    print("Error: Environment variables RIVERDATA - indicating the name of xls river file - must be defined.")
    sys.exit(1)
xlsfile = Path(xlsfile)

meteodir = os.getenv("RIVERMETEODIR")
if meteodir is None:
    print("Error: Environment variables RIVERMETEODIR - indicating where to find discharge files - must be defined.")
    sys.exit(1)
meteodir = Path(meteodir)

wb = load_workbook(filename=xlsfile, read_only=False, data_only=True)
sh_var = wb['biogeochemical_variable']
sh_loc = wb['river_locations']


# read values of variable to be used at river location
NVAR = sh_var.max_row - 1  # 51 + 1
IDvar = np.zeros(NVAR, dtype=int)
#CONCvar=np.zeros(NVAR)
NAMEvar = []
UNITvar = []
for row_index in range(NVAR):
    IDvar[row_index]=  sh_var.cell(row=row_index+2,column=1).value
    NAMEvar.append(str(sh_var.cell(row=row_index+2,column=2).value))
    UNITvar.append(str(sh_var.cell(row=row_index+2,column=3).value))
    #CONCvar[row_index]=sh_var.cell(row_index+1,3).value


bgc_vars = [
    rivers.BGCVar(IDvar[i], NAMEvar[i], UNITvar[i]) for i in range(NVAR)
]


def get_list(f):
    L = []
    for c in f:
        if isinstance(c, tuple):
            v = c[0].value
        else:
            v=c.value  
        if v is not None:
            try:
                a=float(v)
            except:
                a=str(v)
            L.append(a)
    return L
f=sh_loc['B3:B100']

column_list = get_list(f)
nRivers = int(max(column_list))
 

RATIOS = np.zeros((nRivers), np.float32)
for iRiver in range(nRivers):
    irow=4+iRiver
    f=sh_loc[irow]
    row_list= get_list(f)
    R = rivers.River(row_list, meteo_dir=meteodir)
    RATIOS[iRiver] = R.get_ratio()


actual_vs_clim_ratio = np.nanmean(RATIOS)
if np.isnan(actual_vs_clim_ratio): actual_vs_clim_ratio = 1.0
print("Applying ratio : ", actual_vs_clim_ratio)

RIVERS=[]

for iRiver in range(nRivers):
    irow=4+iRiver
    f=sh_loc[irow]
    row_list= get_list(f)
    R = rivers.River(row_list, meteo_dir=meteodir)
    R.apply_ratio(actual_vs_clim_ratio)
    RIVERS.append(R)


def get_RiverPHYS_Data(lato, varname, timelist, Mask):
    return rivers.get_RiverPHYS_Data(lato, varname, timelist, Mask, RIVERS)


def get_RiverBFM_Data(lato, varname):
    return rivers.get_RiverBFM_Data(lato, varname, RIVERS, bgc_vars)


if __name__ == "__main__":
    # print(rivers.save_river_csv(RIVERS, bgc_vars))
    timelist=['20220614-12:00:00']
    from general import mask
    Mask = mask('/g100_work/OGS_prodC/MIT/V1M-dev/V1/devel/wrkdir/BC_IC/mask.nc')
    A=get_RiverPHYS_Data('W','V',timelist,Mask)

